import re
import tkinter as tk
from tkinter import messagebox, filedialog
from ttkbootstrap import Window, Frame, Label, Button, Treeview, Progressbar, constants as ttkconst
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
import PyPDF2
from docx import Document
import openpyxl
from bs4 import BeautifulSoup
import os
import logging
import time
import platform
from concurrent.futures import ThreadPoolExecutor, TimeoutError
from datetime import datetime, timedelta
from collections import defaultdict

# App Constants and Settings
APP_NAME = "Lead Extractor Open"
VERSION = "1.0.0"
LOG_FILE = "extractor_log.txt"
FREE_EMAIL_LIMIT = 100
FREE_PHONE_LIMIT = 100
FREE_FILE_LIMIT = 3

# Setup Logging
logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(message)s")
logging.debug(f"Running on {platform.system()} {platform.release()}")

# Global Variables
license_status = "Free Lifetime Version"
header_label = None
status_label = None
progress_var = None
summary_tree = None

# Extraction Logic
def extract_emails(text):
    pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
    return re.findall(pattern, text)

def extract_phones(text):
    pattern = r"[+]?\d{1,4}?[-.\s\(]?\d{1,3}?\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9}"
    return re.findall(pattern, text)

def clean_and_classify_phone(phone, remove_plus91):
    cleaned = re.sub(r"[ \-\(\)]", "", phone)
    if remove_plus91 and cleaned.startswith("+91"):
        cleaned = cleaned[3:]
    if not cleaned.isdigit():
        return "invalid", phone
    if len(cleaned) == 10 and cleaned[0] in "6789":
        return "mobile", cleaned
    elif len(cleaned) >= 10:
        return "landline", cleaned
    else:
        return "invalid", cleaned

def read_file_content(file_path):
    text = ""
    try:
        if file_path.lower().endswith((".txt", ".csv")):
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        elif file_path.lower().endswith(".pdf"):
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        text += t
        elif file_path.lower().endswith(".docx"):
            doc = Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
        elif file_path.lower().endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for i, sheet in enumerate(wb.sheetnames, 1):
                sheet_data = wb[sheet]
                for row in sheet_data.iter_rows(values_only=True):
                    text += f"[Sheet {i}] " + " ".join(str(c) for c in row if c) + "\n"
        elif file_path.lower().endswith((".html", ".htm")):
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                soup = BeautifulSoup(f, "html.parser")
                text = soup.get_text()
        else:
            logging.warning(f"Unsupported file extension: {os.path.splitext(file_path)[1]}")
    except Exception as e:
        logging.error(f"Error reading file {file_path}: {str(e)}")
    return text

def process_file(file_path, file_types, progress_var, total_files, current_file):
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in file_types:
        logging.warning(f"Unsupported file: {file_path} (extension: {ext})")
        return [], [], [f"Unsupported file: {file_path}"], defaultdict(int), defaultdict(int)
    try:
        with ThreadPoolExecutor() as ex:
            content = ex.submit(read_file_content, file_path).result(timeout=25)
        progress_var.set((current_file / total_files) * 100)
        emails = extract_emails(content)
        phones = extract_phones(content)
        email_sources = defaultdict(int)
        phone_sources = defaultdict(int)
        filename = os.path.basename(file_path)
        for e in emails:
            email_sources[filename] += 1
        for p in phones:
            phone_sources[filename] += 1
        return emails, phones, [], email_sources, phone_sources
    except TimeoutError:
        return [], [], [f"Timeout: {file_path}"], defaultdict(int), defaultdict(int)
    except Exception as e:
        logging.error(f"Error processing {file_path}: {str(e)}")
        return [], [], [f"Error processing {file_path}: {str(e)}"], defaultdict(int), defaultdict(int)

def show_premium_popup():
    message = ("Upgrade to Lead Extractor Pro for unlimited emails, phones, and files!\n"
               "Enjoy email verification, data enrichment, CRM integrations, and more.\n"
               "Begin a 7-day free trial or purchase for $9.99/month or $29 one-time.")
    if messagebox.askyesno("Upgrade to Pro", message, parent=root):
        # Placeholder for trial start (implement URL or payment gateway later)
        ToastNotification("Trial Started", "7-day trial activated! Visit our website for full access.", duration=5000, bootstyle=SUCCESS)

def upload_and_extract():
    global status_label
    extract_emails_flag = True
    extract_phones_flag = True
    remove_plus91 = True
    file_types = [".txt", ".csv", ".pdf", ".docx", ".xlsx", ".xlsm", ".html", ".htm"]

    if not (extract_emails_flag or extract_phones_flag):
        ToastNotification("Warning", "Select at least one: Emails or Phones.", duration=3000, bootstyle=WARNING)
        return None, None

    status_label.configure(text="Processing...")
    progress_var.set(0)
    root.update()

    try:
        root.lift()
        root.focus_force()
        choice = messagebox.askyesno("Choose Input", "Yes = Select Files, No = Select Folder", parent=root)
        logging.debug(f"User chose: {choice}")
    except Exception as e:
        logging.error(f"Error showing input choice: {str(e)}")
        ToastNotification("Error", f"Failed to show input choice: {str(e)}", duration=3000, bootstyle=DANGER)
        status_label.configure(text=f"Error ({license_status})")
        return None, None

    all_emails = []
    mobiles, landlines, invalids = set(), set(), set()
    errors = []
    email_sources, phone_sources = defaultdict(int), defaultdict(int)

    if choice:
        paths = filedialog.askopenfilenames(title="Select Files", parent=root)
        total_files = len(paths)
        if total_files > FREE_FILE_LIMIT:
            paths = paths[:FREE_FILE_LIMIT]
            errors.append(f"Free version: Limited to {FREE_FILE_LIMIT} files.")
        for i, fp in enumerate(paths, 1):
            e, p, er, file_email, file_phone = process_file(fp, file_types, progress_var, total_files, i)
            all_emails.extend([(email, "") for email in e])  # Remove source filename
            for phone in p:
                cat, cleaned = clean_and_classify_phone(phone, remove_plus91)
                if cat == "mobile":
                    mobiles.add((cleaned, ""))
                elif cat == "landline":
                    landlines.add((cleaned, ""))
                else:
                    invalids.add((cleaned, ""))
            errors.extend(er)
            for f, c in file_email.items():
                email_sources[f] += c
            for f, c in file_phone.items():
                phone_sources[f] += c
            root.update()
    else:
        folder = filedialog.askdirectory(title="Select Folder", parent=root)
        if folder:
            all_files = [os.path.join(root_, file) for root_, _, files in os.walk(folder) for file in files]
            total_files = len(all_files)
            for i, fp in enumerate(all_files, 1):
                e, p, er, file_email, file_phone = process_file(fp, file_types, progress_var, total_files, i)
                all_emails.extend([(email, "") for email in e])  # Remove source filename
                for phone in p:
                    cat, cleaned = clean_and_classify_phone(phone, remove_plus91)
                    if cat == "mobile":
                        mobiles.add((cleaned, ""))
                    elif cat == "landline":
                        landlines.add((cleaned, ""))
                    else:
                        invalids.add((cleaned, ""))
                errors.extend(er)
                for f, c in file_email.items():
                    email_sources[f] += c
                for f, c in file_phone.items():
                    phone_sources[f] += c
                root.update()

    if total_files > FREE_FILE_LIMIT or len(all_emails) > FREE_EMAIL_LIMIT or (len(mobiles) + len(landlines) + len(invalids)) > FREE_PHONE_LIMIT:
        all_emails = all_emails[:FREE_EMAIL_LIMIT]
        mobiles = set(list(mobiles)[:FREE_PHONE_LIMIT // 3])
        landlines = set(list(landlines)[:FREE_PHONE_LIMIT // 3])
        invalids = set(list(invalids)[:FREE_PHONE_LIMIT // 3])
        errors.append(f"Free version limits applied: {FREE_EMAIL_LIMIT} emails, {FREE_PHONE_LIMIT} phones, {FREE_FILE_LIMIT} files.")

    # Display summary
    if summary_tree:
        summary_tree.delete(*summary_tree.get_children())
        summary_tree.insert("", "end", values=("Emails", len(all_emails)))
        summary_tree.insert("", "end", values=("Mobiles", len(mobiles)))
        summary_tree.insert("", "end", values=("Landlines", len(landlines)))
        summary_tree.insert("", "end", values=("Invalid Phones", len(invalids)))
        summary_tree.insert("", "end", values=("Errors", len(errors)))

    # Save output
    if any([all_emails, mobiles, landlines, invalids, errors]):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
            title="Save Extracted Data",
            parent=root
        )
        if save_path:
            try:
                if save_path.endswith('.xlsx'):
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    if all_emails:
                        sheet = wb.create_sheet('Emails')
                        sheet.append(["Email"])
                        for email, _ in all_emails:  # Remove source
                            sheet.append([email])
                    if mobiles:
                        sheet = wb.create_sheet('Mobiles')
                        sheet.append(["Number"])
                        for num, _ in mobiles:  # Remove source
                            sheet.append([num])
                    if landlines:
                        sheet = wb.create_sheet('Landlines')
                        sheet.append(["Number"])
                        for num, _ in landlines:  # Remove source
                            sheet.append([num])
                    if invalids:
                        sheet = wb.create_sheet('Invalid Phones')
                        sheet.append(["Number"])
                        for num, _ in invalids:  # Remove source
                            sheet.append([num])
                    if errors:
                        sheet = wb.create_sheet('Errors')
                        sheet.append(["Error"])
                        for error in errors:
                            sheet.append([error])
                    wb.save(save_path)
                else:
                    if all_emails:
                        with open(save_path.replace('.csv', '_emails.csv'), 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerow(["Email"])
                            for email, _ in all_emails:  # Remove source
                                writer.writerow([email])
                            f.close()
                    if mobiles:
                        with open(save_path.replace('.csv', '_mobiles.csv'), 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerow(["Number"])
                            for num, _ in mobiles:  # Remove source
                                writer.writerow([num])
                            f.close()
                    if landlines:
                        with open(save_path.replace('.csv', '_landlines.csv'), 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerow(["Number"])
                            for num, _ in landlines:  # Remove source
                                writer.writerow([num])
                            f.close()
                    if invalids:
                        with open(save_path.replace('.csv', '_invalid_phones.csv'), 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerow(["Number"])
                            for num, _ in invalids:  # Remove source
                                writer.writerow([num])
                            f.close()
                    if errors:
                        with open(save_path.replace('.csv', '_errors.csv'), 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerow(["Error"])
                            for error in errors:
                                writer.writerow([error])
                            f.close()
                ToastNotification("Task Completed", f"Extraction completed: {len(all_emails)} emails, {len(mobiles)} mobiles, {len(landlines)} landlines, {len(invalids)} invalid\nData saved to {save_path}", duration=5000, bootstyle=SUCCESS)
                logging.debug(f"Task completed: Data saved to {save_path}")
            except Exception as e:
                ToastNotification("Save Error", f"Failed to save: {str(e)}", duration=3000, bootstyle=DANGER)
                logging.error(f"Save failed: {str(e)}")
                ToastNotification("Extraction Completed", f"Extraction done: {len(all_emails)} emails, {len(mobiles)} mobiles, {len(landlines)} landlines, {len(invalids)} invalid", duration=4000, bootstyle=SUCCESS)
        else:
            ToastNotification("Extraction Completed", f"Extraction done: {len(all_emails)} emails, {len(mobiles)} mobiles, {len(landlines)} landlines, {len(invalids)} invalid", duration=4000, bootstyle=SUCCESS)

    status_label.configure(text=f"Completed ({license_status})")
    logging.debug(f"Extraction completed with {len(all_emails)} emails, {len(mobiles)} mobiles, {len(landlines)} landlines, {len(invalids)} invalid")
    show_premium_popup()  # Trigger popup after extraction
    return email_sources, phone_sources

# UI
root = Window(themename="flatly")
root.title(f"{APP_NAME} v{VERSION}")
root.geometry("1100x720")

progress_var = tk.DoubleVar()

header = Frame(root, bootstyle="primary")
header.pack(fill=X)
Label(header, text=APP_NAME, font=("Segoe UI", 20, "bold"), bootstyle="inverse-primary").pack(side=LEFT, padx=20, pady=10)
header_label = Label(header, text=f"v{VERSION} — {license_status}", font=("Segoe UI", 10), bootstyle="inverse-primary")
header_label.pack(side=RIGHT, padx=20)

content = Frame(root)
content.pack(fill=BOTH, expand=True)

# Sidebar
sidebar = Frame(content, width=230, bootstyle="dark")
sidebar.pack(side=LEFT, fill=Y)
sidebar.pack_propagate(False)

def sidebar_btn(text, icon, cmd, style="secondary"):
    b = Button(sidebar, text=f"  {icon}  {text}", bootstyle=style, command=cmd)
    b.pack(fill=X, pady=5, padx=10, ipady=8)
    return b

Label(sidebar, text="Navigation", bootstyle="inverse-dark", font=("Segoe UI", 11, "bold")).pack(pady=(15,5))
sidebar_btn("Extract Data", "📂", upload_and_extract)

# Main Content
main_frame = Frame(content)
main_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=10, pady=10)

status_frame = Frame(main_frame)
status_frame.pack(fill=X, pady=5)
status_label = Label(status_frame, text=f"Ready ({license_status})")
status_label.pack(side=LEFT)
progress_bar = Progressbar(status_frame, bootstyle="info-striped", variable=progress_var, maximum=100)
progress_bar.pack(side=LEFT, padx=10, fill=X, expand=True)

Label(main_frame, text="Extraction Progress", font=("Segoe UI", 14, "bold")).pack(anchor=W, pady=10)
progress_bar.pack(fill=X, padx=5, pady=5)

summary_tree = Treeview(main_frame, columns=("Type", "Count"), show="headings", height=10, bootstyle="info")
summary_tree.heading("Type", text="Type")
summary_tree.heading("Count", text="Count")
summary_tree.column("Type", width=250)
summary_tree.column("Count", width=100)
summary_tree.pack(fill=X, pady=5)

root.mainloop()